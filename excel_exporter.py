"""Append extracted rows to an Excel workbook, highlight flagged rows in red."""
from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

import config

log = logging.getLogger(__name__)

HEADERS = [
    "received_at",
    "source",
    "email_subject",
    "vendor_name",
    "invoice_number",
    "invoice_date",
    "due_date",
    "currency",
    "subtotal",
    "tax",
    "total",
    "payment_terms",
    "line_items_json",
    "validation_ok",
    "validation_issues",
    "attachment_file",
]

FLAG_FILL = PatternFill(start_color="FFF8C8C8", end_color="FFF8C8C8", fill_type="solid")
HEADER_FONT = Font(bold=True)


def _open() -> tuple[Workbook, Any]:
    path = config.EXCEL_OUTPUT_PATH
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        ws.append(HEADERS)
        for c in ws[1]:
            c.font = HEADER_FONT
    return wb, ws


def append_rows(rows: list[dict[str, Any]], output_path: Path | None = None) -> Path:
    target = Path(output_path) if output_path else config.EXCEL_OUTPUT_PATH
    if output_path and target != config.EXCEL_OUTPUT_PATH:
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        ws.append(HEADERS)
        for c in ws[1]:
            c.font = HEADER_FONT
    else:
        wb, ws = _open()

    for row in rows:
        line_items = row.get("line_items") or []
        values = [
            row.get("received_at"),
            row.get("source", "gmail"),
            row.get("email_subject", ""),
            row.get("vendor_name"),
            row.get("invoice_number"),
            row.get("invoice_date"),
            row.get("due_date"),
            row.get("currency"),
            row.get("subtotal"),
            row.get("tax"),
            row.get("total"),
            row.get("payment_terms"),
            json.dumps(line_items, ensure_ascii=False),
            row.get("validation_ok"),
            "; ".join(row.get("validation_issues") or []),
            row.get("attachment_file"),
        ]
        ws.append(values)
        if not row.get("validation_ok", True):
            r = ws.max_row
            for c in ws[r]:
                c.fill = FLAG_FILL

    for col_idx, _ in enumerate(HEADERS, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = 22

    wb.save(target)
    log.info("Wrote %d rows to %s", len(rows), target)
    return target
